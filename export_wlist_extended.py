#%%
"""
Export extended wlist fields to CSV and/or a formatted Excel workbook.

- Connects via wlist.package_wlist.get_db_connection()/fetch_dataframe
- Executes the provided SQL (adjusted for minor syntax issues)
- Default output: Excel to /Users/glennehmke/MEGA/Taxonomy/wlist/wlist_extended.xlsx
- Optional: CSV to /Users/glennehmke/MEGA/Taxonomy/wlist/wlist_extended.csv

Run examples:
  python -m wlist.export_wlist_extended                 # Excel only (default)
  python -m wlist.export_wlist_extended --csv           # Excel + CSV
  python -m wlist.export_wlist_extended --no-excel      # CSV only (if --csv given)
"""
#%%
from __future__ import annotations
import os
import sys
import argparse
from typing import Optional

import pandas as pd

# Ensure project root on sys.path when run directly from repository
HERE = os.path.dirname(os.path.abspath(__file__))
PROJ_ROOT = os.path.dirname(HERE)
if PROJ_ROOT not in sys.path:
    sys.path.insert(0, PROJ_ROOT)

# Import DB/CSV helpers (support both package and script execution contexts)
try:
    from . import package_wlist as pw  # type: ignore
except Exception:  # pragma: no cover
    import package_wlist as pw  # type: ignore

# Output paths (absolute, per issue description)
CSV_PATH = "/Users/glennehmke/MEGA/Taxonomy/wlist/wlist_extended.csv"
XLSX_PATH = "/Users/glennehmke/MEGA/Taxonomy/wlist/wlist_extended.xlsx"

# SQL from issue description, with corrections:
# - Added missing comma after alist_change_note
# - Removed erroneous duplicate column reference `wlist.wlist.notes`
# - Kept alias for t_order as "order" (quoted by pandas as-is)
SQL = (
    """
SELECT
  wlist.taxon_sort,
  wlist.is_ultrataxon,
  wlist.taxon_level,
  wlist.sp_id,
  wlist.taxon_id,
  wlist.taxon_name,
  wlist.taxon_name,
  wlist.family_name,
  wlist.family_scientific_name,
  wlist.t_order AS "order",
  wlist.avibase_id,
  wlist.notes,
  wlist.bird_group,
  wlist.population,
  CASE
    WHEN wlist.is_coastal = 1 THEN 'Obligate'
    WHEN wlist.is_coastal = 2 THEN 'Facultative'
  ELSE NULL END AS is_coastal,
  wlist.aust_rli_1990,
  wlist.aust_rli_2000,
  wlist.aust_rli_2010,
  wlist.aust_rli,
  rl.category AS aust_rli_status_desc,
  wlist.alist_change,
  wlist.alist_change_note,
  wlist.is_core,
  wlist.ssp_required,
  wlist.rli_required,
  wlist.reference
FROM wlist
LEFT JOIN lut_rli rl ON wlist.aust_rli = rl.id
ORDER BY wlist.taxon_sort
;
"""
)


def fetch_df() -> pd.DataFrame:
    with pw.get_db_connection() as conn:
        df = pd.read_sql(SQL, conn)
    return df


def fetch_text_columns(table_name: str = "wlist") -> set[str]:
    """Return a set of column names whose DB data_type is TEXT for the given table.
    Falls back to empty set on any error (so normal auto-fit proceeds).
    """
    try:
        with pw.get_db_connection() as conn:
            q = (
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_name = %s
                  AND data_type ILIKE 'text'
                """
            )
            df = pd.read_sql(q, conn, params=(table_name,))
        return set(df["column_name"].str.lower().tolist())
    except Exception:
        return set()


# ---- Excel builder helpers ----

def _estimate_col_width(values, header: str, min_w=8, max_w=60, char_factor=1.1) -> float:
    """Rough width estimator for Excel columns (openpyxl units)."""
    def _len(x):
        # Be robust to array-like values where pd.isna(x) returns an array,
        # which would raise an "ambiguous truth value" error in an if-statement.
        try:
            is_na = pd.isna(x)
            if isinstance(is_na, (bool, type(None))):
                if is_na:
                    return 0
            # If is_na is array-like, ignore and fall through to string conversion.
        except Exception:
            # If pd.isna itself fails on the type, just fall through.
            pass
        try:
            s = "" if x is None else str(x)
        except Exception:
            s = ""
        return len(s)
    lengths = [len(header)] + [_len(v) for v in values]
    p = 0 if not lengths else sorted(lengths)[int(0.90 * (len(lengths) - 1))]
    width = max(min_w, min(max_w, p * char_factor))
    return float(width)


def build_excel(df: pd.DataFrame, dest_path: str) -> None:
    os.makedirs(os.path.dirname(dest_path), exist_ok=True)

    # Write data frame first via pandas -> openpyxl engine
    with pd.ExcelWriter(dest_path, engine="openpyxl") as xlw:
        sheet_name = "wlist_extended"
        # Ensure taxon_id is written as text and preserves blanks
        if "taxon_id" in df.columns:
            df["taxon_id"] = df["taxon_id"].map(lambda x: "" if pd.isna(x) else str(x))
        df.to_excel(xlw, sheet_name=sheet_name, index=False)
        wb = xlw.book
        ws = xlw.sheets[sheet_name]

        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        # Styles
        # Header: dark gray background with white bold text
        header_fill = PatternFill("solid", fgColor="404040")
        header_font = Font(bold=True, color="FFFFFF")
        body_font = Font(bold=False, color="000000")
        body_font_bold = Font(bold=True, color="000000")
        left_align = Alignment(vertical="top", wrap_text=False, horizontal="left")
        # No borders per requirement

        n_rows = ws.max_row
        n_cols = ws.max_column

        # Header row (1-based index)
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = left_align

        # Identify important columns by header name
        headers = [ws.cell(row=1, column=c).value for c in range(1, n_cols + 1)]
        def col_index(name: str) -> Optional[int]:
            try:
                return headers.index(name) + 1
            except ValueError:
                return None

        taxon_level_c = col_index("taxon_level")
        # locate all columns whose header starts with 'taxon_name' (pandas may rename duplicates as 'taxon_name.1')
        taxon_name_cols = [i + 1 for i, h in enumerate(headers) if isinstance(h, str) and h.startswith("taxon_name")]
        taxon_id_c = col_index("taxon_id")

        # Zebra rows for body (alternating white and light gray)
        zebra_a = PatternFill("solid", fgColor="FFFFFF")
        zebra_b = PatternFill("solid", fgColor="F2F2F2")

        for r in range(2, n_rows + 1):
            # Row shading
            fill = zebra_a if (r % 2 == 0) else zebra_b
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = body_font
                cell.fill = fill
                cell.alignment = left_align

            # Bold taxon_name cells when taxon_level == 'sp'
            if taxon_level_c and taxon_name_cols:
                level_val = ws.cell(row=r, column=taxon_level_c).value
                if isinstance(level_val, str) and level_val.strip().lower() == "sp":
                    for c in taxon_name_cols:
                        ws.cell(row=r, column=c).font = body_font_bold

            # Indent if taxon_level == 'ssp' (keep; no wrapping changes)
            if taxon_level_c and taxon_name_cols:
                level_val = ws.cell(row=r, column=taxon_level_c).value
                if isinstance(level_val, str) and level_val.strip().lower() == "ssp":
                    indented_align = Alignment(horizontal="left", vertical="top", indent=2, wrap_text=False)
                    for c in taxon_name_cols:
                        ws.cell(row=r, column=c).alignment = indented_align

        # Ensure taxon_id column cells are text formatted in Excel
        if taxon_id_c:
            for r in range(2, n_rows + 1):
                ws.cell(row=r, column=taxon_id_c).number_format = '@'

        # Column widths — Option B: keep auto-fit for most, exclude TEXT-typed DB columns
        text_cols = fetch_text_columns("wlist")
        DEFAULT_MIN_W, DEFAULT_MAX_W = 8, 60
        TEXT_FIXED_W = 40  # fixed width for TEXT columns to avoid over-wide columns

        for c in range(1, n_cols + 1):
            col_letter = get_column_letter(c)
            header = headers[c - 1] if c - 1 < len(headers) else f"Col{c}"
            series = df[header] if header in df.columns else pd.Series([], dtype="object")

            base_name = header.split(".")[0].lower() if isinstance(header, str) else ""
            if isinstance(header, str) and base_name in text_cols:
                width = float(TEXT_FIXED_W)
            else:
                width = _estimate_col_width(series.values.tolist(), header, min_w=DEFAULT_MIN_W, max_w=DEFAULT_MAX_W)
            ws.column_dimensions[col_letter].width = width

        # Freeze header, set filter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions


def main() -> None:
    parser = argparse.ArgumentParser(description="Export extended wlist data to CSV and/or Excel")
    parser.add_argument("--csv", action="store_true", help="Also export CSV to wlist_extended.csv")
    # Default Excel = True; allow opt-out with --no-excel
    parser.add_argument("--excel", dest="excel", action="store_true", help="Export Excel (default)")
    parser.add_argument("--no-excel", dest="excel", action="store_false", help="Disable Excel export")
    parser.set_defaults(excel=True)
    args = parser.parse_args()

    df = fetch_df()

    did_any = False
    if args.excel:
        build_excel(df, XLSX_PATH)
        print(f"Wrote XLSX: {XLSX_PATH}  (rows: {len(df)})")
        did_any = True

    if args.csv:
        pw.export_csv(df, CSV_PATH)
        print(f"Wrote CSV: {CSV_PATH}  (rows: {len(df)})")
        did_any = True

    if not did_any:
        # Fallback: keep previous default behavior (CSV) if both disabled accidentally
        pw.export_csv(df, CSV_PATH)
        print(f"Wrote CSV (fallback): {CSV_PATH}  (rows: {len(df)})")


if __name__ == "__main__":
    main()
